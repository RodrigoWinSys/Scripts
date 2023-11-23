import os
import pandas as pd
import openpyxl

# Ruta de la carpeta que contiene los archivos de Excel
carpeta = "C:/Users/Rodrigo Rodriguez/OneDrive - Win Systems Group/PBI Data Bases/Market Share/MS 2023"

# Lista los archivos en la carpeta
archivos_en_carpeta = [archivo for archivo in os.listdir(carpeta) if archivo.endswith(".xlsx")]

# Verificar si se encontraron archivos XLSX
if archivos_en_carpeta:
    # Inicializa un DataFrame vacío para almacenar los datos combinados
    datos_combinados = pd.DataFrame()
    archivos_con_problemas = []

    # Itera a través de todos los archivos
    for archivo in archivos_en_carpeta:
        archivo_path = os.path.join(carpeta, archivo)

        try:
            # Abre el archivo Excel utilizando openpyxl
            wb = openpyxl.load_workbook(archivo_path, data_only=True)

            # Itera a través de las hojas no ocultas
            for nombre_hoja in wb.sheetnames:
                ws = wb[nombre_hoja]
                if not ws.sheet_state == 'hidden':
                    # Busca la fila que contiene "DATE"
                    hoja_a_leer = None
                    for fila in ws.iter_rows(min_row=4, max_row=4):
                        for celda in fila:
                            if celda.value == "DATE":
                                hoja_a_leer = nombre_hoja
                                break
                        if hoja_a_leer:
                            break
                    
                    if hoja_a_leer:
                        # Lee el archivo Excel a partir de la fila que contiene "DATE"
                        df = pd.read_excel(archivo_path, sheet_name=hoja_a_leer, header=3)

                        # Elimina la columna "Unnamed" si existe
                        if "Unnamed" in df.columns:
                            df = df.drop(columns=df.filter(like="Unnamed").columns)

                        # Agrega los datos al DataFrame combinado utilizando concat
                        datos_combinados = pd.concat([datos_combinados, df], ignore_index=True)
                        print(f"Archivo agregado con éxito: {archivo_path}")

        except Exception as e:
            print(f"Error al procesar el archivo {archivo_path}: {str(e)}")
            archivos_con_problemas.append(archivo_path)

    # Guarda los datos combinados en un nuevo archivo Excel en la misma carpeta de origen
    archivo_salida = os.path.join(carpeta, "C:/Users/Rodrigo Rodriguez/OneDrive - Win Systems Group/PBI Data Bases/Market Share/datos_combinados.xlsx")
    datos_combinados.to_excel(archivo_salida, index=False)

    # Imprime archivos que causaron problemas
    if archivos_con_problemas:
        print("\nArchivos que no se pudieron agregar:")
        for archivo_problema in archivos_con_problemas:
            print(archivo_problema)
else:
    print("No se encontraron archivos XLSX en la carpeta.")

